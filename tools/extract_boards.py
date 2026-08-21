import ast
import json
import re
import sys
from pathlib import Path

import openpyxl


MODE_NAMES = {
    1: "新手",
    2: "娱乐",
    3: "进阶",
    4: "萌狼",
}


def main():
    if len(sys.argv) < 3:
        raise SystemExit("Usage: python tools/extract_boards.py <input.xlsx> <output.json>")

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    role_map = read_role_map(Path(__file__).resolve().parents[2] / "packages" / "lrs-types" / "src" / "Macro" / "Role" / "RoleId.ts")

    workbook = openpyxl.load_workbook(input_path, data_only=True, read_only=True)
    sheet = workbook["板子配置"] if "板子配置" in workbook.sheetnames else workbook.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [normalize_cell(value) for value in rows[3]]
    header_index = {name: index for index, name in enumerate(headers) if name}

    boards = []
    used_role_ids = set()

    for row in rows[4:]:
        data = {name: normalize_cell(row[index] if index < len(row) else "") for name, index in header_index.items()}
        if not data.get("id") or not data.get("templateName"):
            continue

        board_id = int(float(data["id"]))
        player_count = int(float(data.get("playerCount") or 0))
        mode_type = int(float(data.get("gameModeType") or 0))
        role_pool_ids = parse_number_list(data.get("ROLES", ""))
        option_ids = unique(parse_number_list(data.get("template", "")) + role_pool_ids)

        used_role_ids.update(option_ids)
        used_role_ids.update(role_pool_ids)

        boards.append(
            {
                "id": str(board_id),
                "sourceId": board_id,
                "sortId": int(float(data.get("sortid") or 0)),
                "name": strip_quotes(data.get("templateName", "")),
                "seatCount": player_count,
                "minPeople": int(float(data.get("minPeoples") or 0)),
                "modeType": mode_type,
                "modeName": MODE_NAMES.get(mode_type, f"类型{mode_type}" if mode_type else "未分类"),
                "roleOptions": [role_item(role_id, role_map) for role_id in option_ids],
                "rolePool": [role_item(role_id, role_map) for role_id in role_pool_ids],
            }
        )

    boards.sort(key=lambda item: (item["modeType"], item["seatCount"], item["sortId"], item["sourceId"]))
    config = {
        "source": str(input_path),
        "extractedAt": "2026-08-21",
        "roleOptions": [role_item(role_id, role_map) for role_id in sorted(used_role_ids)],
        "boards": boards,
    }

    output_path.write_text(json.dumps(config, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Extracted {len(boards)} boards and {len(config['roleOptions'])} roles to {output_path}")


def read_role_map(path):
    text = path.read_text(encoding="utf-8")
    role_map = {}
    for match in re.finditer(r"^\s*(\w+)\s*=\s*(\d+),\s*(?://|/\*\*)\s*([^*\n]+)", text, re.MULTILINE):
        role_id = int(match.group(2))
        label = match.group(3).strip(" */")
        if "(" in label:
            label = label.split("(", 1)[0]
        role_map[role_id] = label.strip()
    return role_map


def normalize_cell(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def strip_quotes(value):
    return value.strip().strip('"').strip()


def parse_number_list(value):
    text = normalize_cell(value)
    if not text:
        return []
    try:
        parsed = ast.literal_eval(text)
        if isinstance(parsed, list):
            return [int(float(item)) for item in parsed if str(item).strip()]
    except Exception:
        pass
    return [int(item) for item in re.findall(r"\d+", text)]


def unique(values):
    output = []
    seen = set()
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        output.append(value)
    return output


def role_item(role_id, role_map):
    label = role_map.get(role_id, f"角色{role_id}")
    return {
        "id": role_id,
        "name": label,
        "camp": infer_camp(label),
    }


def infer_camp(label):
    if re.search(r"狼|石像鬼|恶灵|噩梦|蚀日|梦魇", label):
        return "狼人"
    if re.search(r"村民|羊驼|许仙", label):
        return "平民"
    if re.search(r"丘比特|咒狐|暗恋|夜之贵族", label):
        return "特殊"
    return "神职"


if __name__ == "__main__":
    main()
